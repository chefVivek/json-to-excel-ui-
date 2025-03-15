import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
import json
import openpyxl
from openpyxl.styles import PatternFill
import os

def json_to_excel(json_file_path, excel_file_path):
    try:
        # Read JSON Lines file
        data = []
        with open(json_file_path, 'r', encoding='utf-8') as file:
            for line in file:
                if line.strip():
                    data.append(json.loads(line))
        
        # Convert to DataFrame
        df = pd.DataFrame(data)
        df.to_excel(excel_file_path, index=False, engine='openpyxl')
        
        # Load workbook to format
        wb = openpyxl.load_workbook(excel_file_path)
        ws = wb.active
        
        # Format header
        header_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
        for cell in ws[1]:
            cell.fill = header_fill
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
        
        # Save formatted workbook
        wb.save(excel_file_path)
        
        messagebox.showinfo("Success", f"Successfully converted {json_file_path} to {excel_file_path}")
    except FileNotFoundError:
        messagebox.showerror("Error", f"File {json_file_path} not found.")
    except json.JSONDecodeError as e:
        messagebox.showerror("Error", f"Invalid JSON format: {str(e)}")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {str(e)}")

def browse_file():
    file_path = filedialog.askopenfilename(filetypes=[("JSON Files", "*.json")])
    if file_path:
        entry.delete(0, tk.END)
        entry.insert(0, file_path)

def convert_file():
    json_file = entry.get()
    if not json_file or not os.path.exists(json_file):
        messagebox.showerror("Error", "Please select a valid JSON file.")
        return
    excel_file = json_file.replace(".json", ".xlsx")
    json_to_excel(json_file, excel_file)

# Create main Tkinter window
root = tk.Tk()
root.title("JSON to Excel Converter")
root.geometry("400x200")

# Label
label = tk.Label(root, text="Select JSON File:")
label.pack(pady=10)

# Entry field
entry = tk.Entry(root, width=50)
entry.pack()

# Browse button
browse_button = tk.Button(root, text="Browse", command=browse_file)
browse_button.pack(pady=5)

# Convert button
convert_button = tk.Button(root, text="Convert to Excel", command=convert_file)
convert_button.pack(pady=20)

# Run Tkinter event loop
root.mainloop()
